from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Crea un nuovo workbook o apri un file esistente
wb = load_workbook(r"C:\Users\dimartino\Documents\Codice\Python\codici_una_tantum\autocarri_omologazioni_vecchie\prova.xlsx")
ws = wb.active  # Seleziona il foglio attivo

# --- Rimuovi la griglia per tutto il foglio ---
ws.sheet_view.showGridLines = False

# --- Trova l'ultima riga piena ---
ultima_riga = ws.max_row

# --- Applica bordi tratteggiati da A3:G3 fino all'ultima riga ---
# Definisci il bordo tratteggiato
bordo_tratteggiato = Border(
    left=Side(border_style="dotted", color="000000"),
    right=Side(border_style="dotted", color="000000"),
    top=Side(border_style="dotted", color="000000"),
    bottom=Side(border_style="dotted", color="000000"),
)

# Applica il bordo tratteggiato a tutte le celle da A3 a G[ultima_riga]
for row in ws.iter_rows(min_row=3, max_row=ultima_riga, min_col=1, max_col=7):
    for cell in row:
        cell.border = bordo_tratteggiato

# Unisci le celle da A5 ad A9
ws.merge_cells("A1:G1")

# Scrivi "Utilizzatore" nella cella unita
ws["A1"] = "Utilizzatore"

# Applica il grassetto al testo
ws["A1"].font = Font(bold=True)

# Imposta lo sfondo azzurro
ws["A1"].fill = PatternFill(
    start_color="00B0F0", end_color="00B0F0", fill_type="solid"
)

# Centra il testo nella cella unita
ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

# Salva il file Excel
wb.save(r"C:\Users\dimartino\Documents\Codice\Python\codici_una_tantum\autocarri_omologazioni_vecchie\prova.xlsx")
